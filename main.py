import sys
import os
from functools import partial
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QTabWidget, QWidget, QVBoxLayout, QHBoxLayout,
    QTextEdit, QPushButton, QLabel, QFileDialog, QTableWidget, QTableWidgetItem,
    QLineEdit, QGroupBox, QMessageBox, QHeaderView, QSplitter, QSpinBox,
    QScrollArea
)
from PyQt5.QtCore import Qt
from openpyxl import Workbook, load_workbook


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("禾糖脚本工具")
        self.setMinimumSize(1000, 700)

        # 主 TabWidget
        self.tabs = QTabWidget()
        self.setCentralWidget(self.tabs)

        # Tab 1: 行尾加 ^
        self.tab1 = Tab1_AddCaret()
        self.tabs.addTab(self.tab1, "行尾加 。")

        # Tab 2: 表格处理
        self.tab2 = Tab2_TableProcess()
        self.tabs.addTab(self.tab2, "表格处理")


# ======================== Tab 1: 行尾加 ^ ========================
class Tab1_AddCaret(QWidget):
    def __init__(self):
        super().__init__()
        self.segments = []  # 分段后的文本列表
        layout = QHBoxLayout(self)

        # 左侧：输入
        left_layout = QVBoxLayout()
        left_layout.addWidget(QLabel("输入文本："))
        self.input_edit = QTextEdit()
        self.input_edit.setPlaceholderText("在此粘贴或输入多行文本...")
        left_layout.addWidget(self.input_edit)

        # 中间：按钮 + 分段设置
        mid_layout = QVBoxLayout()
        mid_layout.addStretch()
        mid_layout.addWidget(QLabel("每段行数："))
        self.spin_lines = QSpinBox()
        self.spin_lines.setRange(1, 99999)
        self.spin_lines.setValue(150)
        self.spin_lines.setFixedWidth(80)
        mid_layout.addWidget(self.spin_lines)
        self.btn_convert = QPushButton("转换 →")
        self.btn_convert.setFixedWidth(80)
        self.btn_convert.clicked.connect(self.convert)
        mid_layout.addWidget(self.btn_convert)
        mid_layout.addStretch()

        # 右侧：分段输出（ScrollArea 内动态生成）
        right_layout = QVBoxLayout()
        right_layout.addWidget(QLabel("输出结果（分段显示）："))

        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_content = QWidget()
        self.scroll_layout = QVBoxLayout(self.scroll_content)
        self.scroll_layout.setAlignment(Qt.AlignTop)
        self.scroll_area.setWidget(self.scroll_content)
        right_layout.addWidget(self.scroll_area)

        layout.addLayout(left_layout, 1)
        layout.addLayout(mid_layout)
        layout.addLayout(right_layout, 1)

    def convert(self):
        text = self.input_edit.toPlainText()
        if not text.strip():
            return
        lines = text.split('\n')
        result_lines = [line + '。' for line in lines]

        # 按行数分段
        chunk_size = self.spin_lines.value()
        self.segments = []
        for i in range(0, len(result_lines), chunk_size):
            chunk = '\n'.join(result_lines[i:i + chunk_size])
            self.segments.append(chunk)

        # 清空旧的输出区域
        self._clear_scroll_layout()

        # 动态生成每段的文本框和复制按钮
        for idx, seg_text in enumerate(self.segments):
            seg_label = QLabel(f"第 {idx + 1} 段（共 {len(self.segments)} 段）：")
            self.scroll_layout.addWidget(seg_label)

            seg_edit = QTextEdit()
            seg_edit.setReadOnly(True)
            seg_edit.setPlainText(seg_text)
            # 限制高度，避免单段撑满整个区域
            seg_edit.setMinimumHeight(120)
            seg_edit.setMaximumHeight(250)
            self.scroll_layout.addWidget(seg_edit)

            btn_copy = QPushButton(f"复制第 {idx + 1} 段")
            btn_copy.clicked.connect(partial(self.copy_segment, idx))
            self.scroll_layout.addWidget(btn_copy)

        self.scroll_layout.addStretch()

    def copy_segment(self, idx):
        if 0 <= idx < len(self.segments):
            QApplication.clipboard().setText(self.segments[idx])
            QMessageBox.information(self, "提示", f"第 {idx + 1} 段已复制到剪贴板")

    def _clear_scroll_layout(self):
        """清空 scroll_layout 中的所有子控件"""
        while self.scroll_layout.count():
            item = self.scroll_layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()


# ======================== Tab 2: 表格处理 ========================
class Tab2_TableProcess(QWidget):
    def __init__(self):
        super().__init__()
        self.split_data = []  # 拆分后的数据: [(序号, 配音角色, 配音内容), ...]
        self.init_ui()

    def init_ui(self):
        main_layout = QVBoxLayout(self)

        # ---- 步骤 1：导入并拆分表格 ----
        group1 = QGroupBox("步骤 1：导入并拆分表格")
        g1_layout = QVBoxLayout(group1)

        file_row = QHBoxLayout()
        self.file_path_label = QLineEdit()
        self.file_path_label.setReadOnly(True)
        self.file_path_label.setPlaceholderText("请选择 .xlsx 文件...")
        file_row.addWidget(self.file_path_label)
        self.btn_select_file = QPushButton("选择文件")
        self.btn_select_file.clicked.connect(self.select_xlsx)
        file_row.addWidget(self.btn_select_file)
        self.btn_split = QPushButton("拆分处理")
        self.btn_split.clicked.connect(self.split_table)
        file_row.addWidget(self.btn_split)
        g1_layout.addLayout(file_row)

        self.split_info_label = QLabel("")
        g1_layout.addWidget(self.split_info_label)

        main_layout.addWidget(group1)

        # ---- 步骤 2：导入提示词 ----
        group2 = QGroupBox("步骤 2：导入提示词")
        g2_layout = QHBoxLayout(group2)

        # 图片提示词（左）
        img_layout = QVBoxLayout()
        img_header = QHBoxLayout()
        img_header.addWidget(QLabel("图片提示词（每行一个）："))
        self.btn_import_img = QPushButton("从文件导入")
        self.btn_import_img.clicked.connect(lambda: self.import_prompt_file(self.img_prompt_edit))
        img_header.addWidget(self.btn_import_img)
        img_layout.addLayout(img_header)
        self.img_prompt_edit = QTextEdit()
        self.img_prompt_edit.setPlaceholderText("粘贴图片提示词，每行一个...")
        img_layout.addWidget(self.img_prompt_edit)

        # 视频提示词（右）
        vid_layout = QVBoxLayout()
        vid_header = QHBoxLayout()
        vid_header.addWidget(QLabel("视频提示词（每行一个）："))
        self.btn_import_vid = QPushButton("从文件导入")
        self.btn_import_vid.clicked.connect(lambda: self.import_prompt_file(self.vid_prompt_edit))
        vid_header.addWidget(self.btn_import_vid)
        vid_layout.addLayout(vid_header)
        self.vid_prompt_edit = QTextEdit()
        self.vid_prompt_edit.setPlaceholderText("粘贴视频提示词，每行一个...")
        vid_layout.addWidget(self.vid_prompt_edit)

        g2_layout.addLayout(img_layout)
        g2_layout.addLayout(vid_layout)
        main_layout.addWidget(group2)

        # ---- 步骤 3：角色匹配 ----
        group3 = QGroupBox("步骤 3：角色匹配")
        g3_layout = QHBoxLayout(group3)
        g3_layout.addWidget(QLabel("角色名称（逗号分隔）："))
        self.role_input = QLineEdit()
        self.role_input.setPlaceholderText("例如：苏晚,谢南辞,江柔")
        g3_layout.addWidget(self.role_input)
        main_layout.addWidget(group3)

        # ---- 步骤 4：合并预览 + 导出 ----
        group4 = QGroupBox("步骤 4：合并预览 + 导出")
        g4_layout = QVBoxLayout(group4)

        btn_row = QHBoxLayout()
        self.btn_merge = QPushButton("合并预览")
        self.btn_merge.clicked.connect(self.merge_preview)
        btn_row.addWidget(self.btn_merge)
        self.btn_export = QPushButton("导出 xlsx")
        self.btn_export.clicked.connect(self.export_xlsx)
        btn_row.addWidget(self.btn_export)
        g4_layout.addLayout(btn_row)

        self.result_table = QTableWidget()
        self.result_table.setColumnCount(9)
        self.result_table.setHorizontalHeaderLabels([
            "镜头ID", "配音角色", "配音内容", "场景",
            "出场角色", "图片提示词", "视频提示词", "情感", "强度"
        ])
        self.result_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.result_table.setEditTriggers(QTableWidget.NoEditTriggers)
        g4_layout.addWidget(self.result_table)

        main_layout.addWidget(group4, 1)  # 让表格区域占更多空间

    # ---- 步骤 1 相关方法 ----
    def select_xlsx(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择 xlsx 文件", "", "Excel 文件 (*.xlsx)"
        )
        if path:
            self.file_path_label.setText(path)

    def split_table(self):
        path = self.file_path_label.text().strip()
        if not path or not os.path.exists(path):
            QMessageBox.warning(self, "提示", "请先选择有效的 xlsx 文件")
            return

        try:
            wb = load_workbook(path)
            ws = wb.active
        except Exception as e:
            QMessageBox.critical(self, "错误", f"读取文件失败：{e}")
            return

        # 第一步：读取所有行，构建逐字符的角色标记
        # 把所有文本融合成一个大字符串，同时记录每个字符属于哪个角色
        char_roles = []  # [(字符, 角色), ...]

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            role = row[1] if len(row) > 1 else None
            text = row[7] if len(row) > 7 else None

            if not text:
                continue

            if not role or str(role).strip() == "":
                role = "旁白"
            else:
                role = str(role).strip()

            # 清洗文本：去掉换行，把 ^ 替换为 。
            text_str = str(text).replace('\r\n', '').replace('\r', '').replace('\n', '')
            text_str = text_str.replace('^', '。')

            for ch in text_str:
                char_roles.append((ch, role))

        # 第二步：按 。分割成分镜段落
        segments = []       # [[(字符, 角色), ...], ...]
        current_seg = []
        for ch, role in char_roles:
            if ch == '。':
                if current_seg:
                    segments.append(current_seg)
                    current_seg = []
            else:
                current_seg.append((ch, role))
        if current_seg:
            segments.append(current_seg)

        # 第三步：每个段落分配镜头ID，按角色分组
        # 一个段落可能跨多个角色 → 同一镜头ID，多行
        self.split_data = []  # [(镜头ID, 配音角色, 配音内容), ...]

        for shot_id, seg_chars in enumerate(segments, 1):
            # 将连续相同角色的字符合并
            groups = []
            cur_role = None
            cur_text = []
            for ch, role in seg_chars:
                if role != cur_role:
                    if cur_text:
                        groups.append((cur_role, ''.join(cur_text).strip()))
                    cur_role = role
                    cur_text = [ch]
                else:
                    cur_text.append(ch)
            if cur_text:
                groups.append((cur_role, ''.join(cur_text).strip()))

            for role, text in groups:
                if text:
                    self.split_data.append((shot_id, role, text))

        unique_shots = len(set(s[0] for s in self.split_data))
        self.split_info_label.setText(
            f"拆分完成，共 {unique_shots} 个分镜，{len(self.split_data)} 行数据"
        )
        QMessageBox.information(
            self, "提示",
            f"拆分完成！共 {unique_shots} 个分镜，{len(self.split_data)} 行数据"
        )

    # ---- 步骤 2 相关方法 ----
    def import_prompt_file(self, target_edit):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择提示词文件", "", "文本文件 (*.txt);;所有文件 (*)"
        )
        if path:
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    content = f.read()
                target_edit.setPlainText(content)
            except Exception as e:
                QMessageBox.critical(self, "错误", f"读取文件失败：{e}")

    # ---- 步骤 3 + 4 合并预览 ----
    def merge_preview(self):
        if not self.split_data:
            QMessageBox.warning(self, "提示", "请先完成步骤 1（拆分表格）")
            return

        # 获取提示词（按镜头ID对应，每个镜头一行提示词）
        img_lines = self.img_prompt_edit.toPlainText().split('\n')
        vid_lines = self.vid_prompt_edit.toPlainText().split('\n')

        # 获取角色名列表
        role_text = self.role_input.text().strip()
        role_names = [r.strip() for r in role_text.split(',') if r.strip()] if role_text else []

        # 构建最终数据
        # 提示词按镜头ID索引（镜头ID从1开始，提示词第0行对应镜头1）
        self.final_data = []
        for i, (shot_id, role, text) in enumerate(self.split_data):
            prompt_idx = shot_id - 1  # 镜头ID从1开始，提示词从0开始
            img_prompt = img_lines[prompt_idx].strip() if prompt_idx < len(img_lines) else ""
            vid_prompt = vid_lines[prompt_idx].strip() if prompt_idx < len(vid_lines) else ""

            # 角色匹配：扫描图片提示词和视频提示词
            matched_roles = []
            for rn in role_names:
                combined = img_prompt + " " + vid_prompt
                if rn in combined:
                    matched_roles.append(rn)
            matched_str = ",".join(matched_roles)

            self.final_data.append((
                str(shot_id),   # 镜头ID
                role,           # 配音角色
                text,           # 配音内容
                "",             # 场景（留空）
                matched_str,    # 出场角色
                img_prompt,     # 图片提示词
                vid_prompt,     # 视频提示词
                "",             # 情感（留空）
                ""              # 强度（留空）
            ))

        # 更新表格
        self.result_table.setRowCount(len(self.final_data))
        for row_idx, row_data in enumerate(self.final_data):
            for col_idx, value in enumerate(row_data):
                item = QTableWidgetItem(value)
                self.result_table.setItem(row_idx, col_idx, item)

        QMessageBox.information(self, "提示", f"合并完成，共 {len(self.final_data)} 行")

    # ---- 步骤 4 导出 ----
    def export_xlsx(self):
        if not hasattr(self, 'final_data') or not self.final_data:
            QMessageBox.warning(self, "提示", "请先完成合并预览")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "导出文件", "镜头列表.xlsx", "Excel 文件 (*.xlsx)"
        )
        if not path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "镜头列表"

            # 写表头
            headers = ["镜头ID", "配音角色", "配音内容", "场景",
                        "出场角色", "图片提示词", "视频提示词", "情感", "强度"]
            ws.append(headers)

            # 写数据
            for row_data in self.final_data:
                ws.append(list(row_data))

            wb.save(path)
            QMessageBox.information(self, "提示", f"导出成功！\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败：{e}")


# ======================== 入口 ========================
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
